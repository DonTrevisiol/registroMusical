# ./RegistroLaboral/exportar_xlsx.py

import os

from tkinter import messagebox

from openpyxl import Workbook

from openpyxl.styles import (
PatternFill,
Font,
Alignment,
Border,
Side
)

from openpyxl.utils import (
get_column_letter
)

from constantes import MESES
from tabla_utils import obtener_datos_fila

def exportar_tabla_mensual_xlsx(
    anio,
    mes,
    registros
):

    carpeta_anio = f"XLSXs/{anio}"

    os.makedirs(
        carpeta_anio,
        exist_ok=True
    )

    nombre_mes = (
        MESES[mes]
        .lower()
    )

    archivo = (
        f"{carpeta_anio}/{nombre_mes}.xlsx"
    )

    wb = Workbook()

    hoja = wb.active

    hoja.title = "Tabla mensual"

    encabezados = [
        "#",
        "Fecha",
        "Canciones",
        "Tiempo total",
        "Tiempo promedio",
        "Dinero total",
        "Dinero promedio",
        "Detalles"
    ]

    hoja.append(encabezados)

    # -------------------------
    # Encabezados
    # -------------------------

    color_encabezado = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    fuente_encabezado = Font(
        color="FFFFFF",
        bold=True
    )

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for celda in hoja[1]:

        celda.fill = color_encabezado

        celda.font = fuente_encabezado

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        celda.border = borde

    # -------------------------
    # Colores de texto
    # -------------------------

    colores_texto = {
        "verde": "008000",
        "rojo": "C00000",
        "azul": "0000FF",
        "morado": "800080",
        "gris": "808080"
    }

    fila_excel = 2

    for registro in registros:

        color, valores = (
            obtener_datos_fila(
                registro
            )
        )

        hoja.append(
            list(valores)
        )

        for columna in range(1, 9):

            celda = hoja.cell(
                row=fila_excel,
                column=columna
            )

            celda.border = borde

            if color in colores_texto:

                celda.font = Font(
                    color=colores_texto[color]
                )

            if columna == 8:

                celda.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

            else:

                celda.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        fila_excel += 1

    # -------------------------
    # Ajustar columnas
    # -------------------------

    for columna in hoja.columns:

        longitud_maxima = 0

        letra_columna = (
            get_column_letter(
                columna[0].column
            )
        )

        for celda in columna:

            try:

                longitud_maxima = max(
                    longitud_maxima,
                    len(str(celda.value))
                )

            except:

                pass

        hoja.column_dimensions[
            letra_columna
        ].width = longitud_maxima + 3

    # -------------------------
    # Vista e impresión
    # -------------------------

    hoja.freeze_panes = "A2"

    hoja.sheet_view.zoomScale = 100

    hoja.page_setup.orientation = "landscape"

    hoja.page_setup.fitToWidth = 1

    hoja.page_setup.fitToHeight = 1

    hoja.sheet_properties.pageSetUpPr.fitToPage = True

    hoja.page_margins.left = 0.2
    hoja.page_margins.right = 0.2
    hoja.page_margins.top = 0.3
    hoja.page_margins.bottom = 0.3

    wb.save(archivo)

    messagebox.showinfo(
        "Archivo exportado",
        f"Tabla guardada en:\n\n{archivo}"
    )
